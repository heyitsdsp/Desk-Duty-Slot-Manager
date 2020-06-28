import os
import random
import pandas as pd
import Whatsapp

# Starting up Whatsapp web 
Whatsapp.initiate()

# Parsing the csv file (Make sure you have everyone's contact saved)
name_file = pd.read_csv('Names.csv')
name_sheet = (name_file.iloc[:, 0]).tolist()
whatsapp_sheet = (name_file.iloc[:, 1]).tolist()

# Sending introduction message to everyone
Whatsapp.send_intro_message(whatsapp_sheet)

days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# importing the modules required to read and modify excel sheets

try:
    import xlrd as xlr
except ImportError:
    print("This file needs XLRD to run , please execute 'pip install xlrd' and run this file again")
    exit()

try:
    import openpyxl
except ImportError:
    print("This file needs Openpyxl to run , please execute 'pip install openpyxl' and run this file again")
    exit()


loc = os.getcwd() + r'/slots.xlsx'

wb = xlr.open_workbook(loc)
new_wb = openpyxl.load_workbook('Template.xlsx')
sheet_names = new_wb.sheetnames
names = []
day = {}

c = 0

while(c!=5):
    new_sheet = new_wb[sheet_names[c]]
    rval = cval = 1

    sheet = wb.sheet_by_index(c)
    for j in range(1,sheet.ncols):
        names.clear()
        for i in range(2,sheet.nrows):
            if(sheet.cell_value(i,j) == 'yes' or 
               sheet.cell_value(i,j)=='Yes'):
                names.append(sheet.cell_value(i,0))
            day[sheet.cell_value(1,j)]=names[0:len(names)]
    tt = {}
    sjt = {}

# Writing the time slots in the "test.xlsx" sheet

    for i in day.keys():
        rval = 1
        tt[i] = []
        sjt[i] = []
        new_sheet.cell(row = rval, column = cval).value = i
        rval+= 5
        new_sheet.cell(row = rval, column = cval).value = i
        cval+= 1

    name = []

# Setting up everyone's slots for TT desk duties

    cval = 1
    
    for i in day.keys():
        rval = 2
        x = 0
        name.clear()     # get rid of redundant data (if any)
        
        if(len(day[i])>3):
            while(len(name) != 3):
                value1 = random.randrange(0, len(day[i])-1)
                names = day[i]
                name.append(names[value1])
                day[i].remove(names[value1])
                
            tt[i] = name[0:len(name)]
        else:
            tt[i] = day[i][1]
            
        # some of the elements turn out to be strings, so, converting
        # them to a list with a single element to prevent bad data
        # from showing up in "test.xlsx"
        
        if(not isinstance(tt[i], list)):
            tt[i] = [tt[i]]

        while(x != len(tt[i])):
            new_sheet.cell(row = rval, column = cval).value = tt[i][x]
            x+= 1
            rval+= 1
        cval+= 1

# Setting up everyone's slots for SJT desk duties
   
    cval = 1
    for i in day.keys():
        rval = 7
        x = 0
        if(len(day[i]) >= 3):
            sjt[i] = day[i][:3]
        else:
            sjt[i] = day[i]
            
        while(x != len(sjt[i])):
            new_sheet.cell(row = rval, column = cval).value = sjt[i][x]
            x+= 1
            rval+= 1
        cval+= 1
    
    
    # For testing purposes, print the dictionaries for tt and sjt
    # Just un-comment the block of code below
# =============================================================================
#     print(tt)
#     print("\n")
#     print(sjt)
#     print('\n')
# =============================================================================
    
    
# Sending Whatsapp messages to everyone in the list
    
# For TT desk duties
    for i in tt:
        for j in tt[i]:
            """
            Introducing a try-except block to deal with an 
            annoying " " that appears after some names in the
            list
            """
            try:
                ind = name_sheet.index(j)
            except:
                j = j[0:len(j)-1]
                ind = name_sheet.index(j)
            
            whatsapp_name = whatsapp_sheet[ind]
            Whatsapp.send_slots_message(whatsapp_name, days[c], i, 'TT')
            
# For SJT desk duties
    for i in sjt:
        for j in sjt[i]:
            """
            Introducing a try-except block to deal with an 
            annoying " " that appears after some names in the
            list
            """
            try:
                ind = name_sheet.index(j)
            except:
                j = j[0:len(j)-1]
                ind = name_sheet.index(j)
            
            whatsapp_name = whatsapp_sheet[ind]
            Whatsapp.send_slots_message(whatsapp_name, days[c], i, 'SJT')
            
    c+=1

# Saving the new modified excel sheet containing slot details for 
# the next week

new_wb.save('test.xlsx')


